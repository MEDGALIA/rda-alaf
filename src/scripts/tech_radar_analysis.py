"""Read-only analysis of the VANTAGE Technology Radar workbook.

Never opens the xlsx in write mode and never modifies any cell. Reports:
  - contiguous ranges of fully-empty rows
  - cells whose type/shape doesn't match what their column expects
  - column coverage: data columns missing from the Dictionary tab, and
    documented columns no data tab actually has
  - vocabulary coverage: terms used but not in the Vocabulary tab, and
    documented terms never used
  - ontology coverage: vocabulary terms not backed by a declared standard
  - review candidates: near-duplicate terms that were deliberately NOT
    auto-merged, so a curator can decide

Nothing is corrected: guessing at real WG records could be worse than
leaving them alone. Which columns get vocabulary checks is driven by the
Dictionary tab's Value Type, so free-text columns aren't reported as
having thousands of "undocumented terms".

Usage:
    python tech_radar_analysis.py [--xlsx PATH] [--out data/reports/workbook_analysis.md]
"""

from __future__ import annotations

import argparse
import datetime
from pathlib import Path

import openpyxl

from radar_sync_common import (
    CONTROLLED_TYPES,
    METADATA_SHEETS,
    iter_header,
    load_workbook_schema,
    row_is_empty,
    split_multi_value,
)

DEFAULT_XLSX = Path("data/VANTAGE-Technology-Radar.xlsx")
DEFAULT_OUT = Path("data/reports/workbook_analysis.md")

NAME_LIKE_HEADERS = {"verified by", "source organization", "resource name"}
DATE_EXPECTED_HEADERS = {"last verified", "deprecation date", "publication date", "discovery date"}
VERSION_HEADERS = {"version"}
URL_HEADERS = {"url"}


def is_date_cell(value) -> bool:
    return isinstance(value, (datetime.date, datetime.datetime))


def is_numeric_cell(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def check_anomaly(header: str, value) -> str | None:
    """Return a description of the anomaly, or None if the cell looks fine."""
    if value is None or str(value).strip() == "":
        return None
    header_lc = header.strip().lower()

    if header_lc in NAME_LIKE_HEADERS:
        if is_date_cell(value):
            return f"'{header}' holds a date, expected a name/organization/title"
        if is_numeric_cell(value):
            return f"'{header}' holds a number ({value!r}), expected a name/organization/title"
    elif header_lc in DATE_EXPECTED_HEADERS:
        if not is_date_cell(value):
            return f"'{header}' is not a real date cell (stored as {type(value).__name__}: {value!r})"
    elif header_lc in VERSION_HEADERS:
        if is_date_cell(value):
            return f"'{header}' holds a date ({value!r}), expected a version string"
    elif header_lc in URL_HEADERS:
        text = str(value).strip()
        if not text.startswith(("http://", "https://")):
            return f"'{header}' does not look like a URL: {value!r}"
    return None


def format_row_ranges(rows: list[int]) -> str:
    if not rows:
        return "none"
    ranges = []
    start = prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        ranges.append((start, prev))
        start = prev = r
    ranges.append((start, prev))
    return ", ".join(f"{a}" if a == b else f"{a}–{b}" for a, b in ranges)


def scan_sheet(ws) -> dict:
    columns = list(iter_header(ws))
    max_row = ws.max_row

    empty_rows = []
    populated_rows = 0
    anomalies = []

    for row_idx in range(2, max_row + 1):
        if row_is_empty(ws, row_idx, columns):
            empty_rows.append(row_idx)
            continue
        populated_rows += 1
        for col_letter, header in columns:
            value = ws[f"{col_letter}{row_idx}"].value
            reason = check_anomaly(header, value)
            if reason:
                anomalies.append((row_idx, header, value, reason))

    return {
        "headers": [h for _c, h in columns],
        "total_rows": max_row - 1,
        "empty_rows": empty_rows,
        "populated_rows": populated_rows,
        "anomalies": anomalies,
    }


def collect_used_terms(wb, schema: dict, data_sheets: list[str]) -> dict[str, dict[str, int]]:
    """For each controlled column, count how often each term is used."""
    used: dict[str, dict[str, int]] = {}
    for col in schema["columns"]:
        if col["value_type"] not in CONTROLLED_TYPES:
            continue
        counts: dict[str, int] = {}
        for sheet_name in data_sheets:
            ws = wb[sheet_name]
            headers = {h: c for c, h in iter_header(ws)}
            if col["name"] not in headers:
                continue
            letter = headers[col["name"]]
            for row_idx in range(2, ws.max_row + 1):
                raw = ws[f"{letter}{row_idx}"].value
                if raw is None or not str(raw).strip():
                    continue
                terms = (
                    split_multi_value(raw, col["separator"] or ";")
                    if col["value_type"] == "controlled_multi"
                    else [str(raw).strip()]
                )
                for t in terms:
                    counts[t] = counts.get(t, 0) + 1
        used[col["name"]] = counts
    return used


def find_missing_controlled(wb, schema: dict, data_sheets: list[str]) -> dict[str, list[tuple]]:
    """Populated rows with an empty controlled-vocabulary cell.

    A blank in a controlled column is usually an oversight rather than a
    deliberate 'not applicable', so it's worth surfacing for curation.
    """
    missing: dict[str, list[tuple]] = {}
    for sheet_name in data_sheets:
        ws = wb[sheet_name]
        headers = {h: c for c, h in iter_header(ws)}
        key_letter = next(iter(headers.values()))
        found = []
        for col in schema["columns"]:
            if col["value_type"] not in CONTROLLED_TYPES or col["name"] not in headers:
                continue
            letter = headers[col["name"]]
            for row_idx in range(2, ws.max_row + 1):
                if row_is_empty(ws, row_idx, list(iter_header(ws))):
                    continue
                value = ws[f"{letter}{row_idx}"].value
                if value is None or not str(value).strip():
                    found.append((row_idx, ws[f"{key_letter}{row_idx}"].value, col["name"]))
        if found:
            missing[sheet_name] = sorted(found)
    return missing


def _word_set(term: str) -> frozenset[str]:
    """Crudely singularized word set, for near-duplicate detection."""
    words = [w.strip("&/,()").lower() for w in term.split()]
    return frozenset(w.rstrip("s") for w in words if w)


def find_review_candidates(terms: list[str]) -> list[tuple[str, str]]:
    """Pairs where one term's words are a subset of another's.

    Catches 'Retrieval' vs 'Retrieval Agents' without flagging genuinely
    distinct siblings like EDAM's several 'Data ...' terms.
    """
    pairs = []
    sets = {t: _word_set(t) for t in terms}
    for i, a in enumerate(terms):
        for b in terms[i + 1 :]:
            if sets[a] and sets[b] and sets[a] != sets[b] and (sets[a] < sets[b] or sets[b] < sets[a]):
                shorter, longer = (a, b) if sets[a] < sets[b] else (b, a)
                pairs.append((shorter, longer))
    return pairs


def render_report(
    xlsx_path: Path,
    schema: dict,
    sheet_results: dict,
    used: dict,
    data_sheets: list[str],
    missing_controlled: dict,
) -> str:
    L = [
        "# VANTAGE Technology Radar — Workbook Analysis",
        "",
        f"Source: `{xlsx_path.as_posix()}`",
        "",
        "Read-only scan. No cell in the workbook was modified. Findings are flagged, not "
        "corrected — fixes are a human judgment call.",
        "",
        "## Per-sheet row counts and cell anomalies",
        "",
    ]
    for sheet_name, result in sheet_results.items():
        L += [
            f"### {sheet_name}",
            "",
            f"- Total data rows: {result['total_rows']}",
            f"- Empty rows: {len(result['empty_rows'])} (rows {format_row_ranges(result['empty_rows'])})",
            f"- Populated rows: {result['populated_rows']}",
            f"- Anomalous cells: {len(result['anomalies'])}",
            "",
        ]
        if result["anomalies"]:
            L += ["| Row | Column | Value | Issue |", "|---|---|---|---|"]
            for row_idx, header, value, reason in result["anomalies"]:
                L.append(f"| {row_idx} | {header} | `{value!r}` | {reason} |")
            L.append("")

    # ---- column coverage
    documented = {c["name"] for c in schema["columns"]}
    actual = {h for r in sheet_results.values() for h in r["headers"]}
    L += ["## Column coverage", ""]
    missing = sorted(actual - documented)
    orphan = sorted(documented - actual)
    L.append(
        f"- Columns in data tabs but **not documented** in `Dictionary`: "
        + (", ".join(f"`{c}`" for c in missing) if missing else "none")
    )
    L.append(
        f"- Documented in `Dictionary` but **present in no data tab**: "
        + (", ".join(f"`{c}`" for c in orphan) if orphan else "none")
    )
    L.append("")

    # ---- vocabulary coverage
    L += [
        "## Vocabulary coverage",
        "",
        "Only columns whose `Value Type` is a controlled vocabulary are checked; free-text "
        "columns are skipped by design.",
        "",
    ]
    for col in schema["columns"]:
        if col["value_type"] not in CONTROLLED_TYPES:
            continue
        vocab = {t["name"] for t in col["terms"]}
        vocab_lc = {v.lower() for v in vocab}
        counts = used.get(col["name"], {})
        undocumented = sorted(t for t in counts if t.lower() not in vocab_lc)
        never_used = sorted(v for v in vocab if v.lower() not in {u.lower() for u in counts})
        L += [
            f"### {col['name']} (`{col['value_type']}`)",
            "",
            f"- Vocabulary size: {len(vocab)}; distinct values in use: {len(counts)}",
            "- Used but **not in `Vocabulary`**: "
            + (", ".join(f"`{t}`" for t in undocumented) if undocumented else "none"),
            "- In `Vocabulary` but **never used**: "
            + (", ".join(f"`{t}`" for t in never_used) if never_used else "none"),
            "",
        ]

    # ---- missing controlled values
    L += ["## Missing controlled values", ""]
    if missing_controlled:
        L += ["Populated rows with an empty controlled-vocabulary cell.", ""]
        for sheet_name, rows in missing_controlled.items():
            L.append(f"**{sheet_name}**")
            L += [f"- row {r}: `{name}` is missing `{col}`" for r, name, col in rows]
            L.append("")
    else:
        L += ["None — every populated row has a value in each controlled column that applies to it.", ""]

    # ---- ontology coverage
    L += [
        "## Ontology coverage",
        "",
        "Declared standards: " + ", ".join(f"`{s['id']}` ({s['url']})" for s in schema["standards"]),
        "",
    ]
    for col in schema["columns"]:
        if col["value_type"] not in CONTROLLED_TYPES:
            continue
        unbacked = sorted(t["name"] for t in col["terms"] if not t["ontology"])
        backed = sorted(f"{t['name']} → {t['ontology']}" for t in col["terms"] if t["ontology"])
        L += [
            f"### {col['name']}",
            "",
            f"- Standards-backed terms ({len(backed)}): " + (", ".join(f"`{b}`" for b in backed) if backed else "none"),
            f"- **Not backed by any standard** ({len(unbacked)}): "
            + (", ".join(f"`{u}`" for u in unbacked) if unbacked else "none"),
            "",
        ]

    # ---- review candidates
    L += [
        "## Review candidates (near-duplicates, deliberately not auto-merged)",
        "",
        "Term pairs where one term's words are a subset of the other's. These were left alone "
        "because merging them is a semantic judgment, not a spelling fix.",
        "",
    ]
    any_pairs = False
    for col in schema["columns"]:
        if col["value_type"] not in CONTROLLED_TYPES:
            continue
        pairs = find_review_candidates([t["name"] for t in col["terms"]])
        if not pairs:
            continue
        any_pairs = True
        L.append(f"**{col['name']}**")
        L += [f"- `{a}` vs `{b}`" for a, b in pairs]
        L.append("")
    if not any_pairs:
        L += ["none", ""]

    return "\n".join(L).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze the Tech Radar workbook.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    before_bytes = args.xlsx.read_bytes()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    schema = load_workbook_schema(wb)
    data_sheets = [s for s in wb.sheetnames if s not in METADATA_SHEETS]
    sheet_results = {name: scan_sheet(wb[name]) for name in data_sheets}
    used = collect_used_terms(wb, schema, data_sheets)
    missing_controlled = find_missing_controlled(wb, schema, data_sheets)
    wb.close()

    after_bytes = args.xlsx.read_bytes()
    assert before_bytes == after_bytes, "xlsx bytes changed during a read-only scan"

    report = render_report(args.xlsx, schema, sheet_results, used, data_sheets, missing_controlled)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(report, encoding="utf-8", newline="\n")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()

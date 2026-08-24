"""Stamp Verified By/Last Verified from a PR's approval, once merged.

Runs after a PR merges (pull_request: closed, merged == true) and only if
it carries an APPROVED review -- approving a PR is meant to be the act of
verifying, but nothing wrote that into the data until now.

Writes directly into a checkout of `main`. This is safe against `main`'s
branch protection because the caller authenticates as the `vt-radar-
verification-bot` GitHub App, which is named in `main`'s
`bypass_pull_request_allowances` -- the built-in Actions app can't be
granted that (tested: accepted with HTTP 200, then silently dropped), only
a self-created GitHub App can. See drafts/VANTAGE-Tech-Radar-Sync-Plan.md,
Step 4.

Only rows whose content actually changed in this PR, and whose
verification is currently blank, get stamped -- a row the editor already
self-attested (fully supplied + freshly touched, xlsx_to_json.py's
existing trust path) keeps that specific claim rather than being
overwritten by a generic approver stamp.

Usage:
    python stamp_verification.py --base-dir DIR --write-dir DIR --xlsx PATH
        --repo owner/repo --pr-number N
"""

from __future__ import annotations

import argparse
import datetime
import json
import subprocess
from pathlib import Path

from radar_sync_common import content_hash, load_sheet_json, save_sheet_json


def _load_records(json_dir: Path) -> dict[str, tuple[str, dict]]:
    """{id: (sheet_stem, record)} across every data sheet json in a dir."""
    records = {}
    for path in json_dir.glob("*.json"):
        if path.stem == "dictionary":
            continue
        data = load_sheet_json(path)
        if not data:
            continue
        for r in data.get("records", []):
            if r.get("id"):
                records[r["id"]] = (path.stem, r)
    return records


def find_changed_row_ids(base_dir: Path, head_dir: Path) -> set[str]:
    """IDs whose substantive content differs between base and head.

    New rows (no base counterpart) count as changed. Verification-only
    edits do not: content_hash() already excludes verified_by/last_verified,
    the same exclusion xlsx_to_json.py relies on elsewhere.
    """
    base = _load_records(base_dir)
    head = _load_records(head_dir)
    changed = set()
    for row_id, (_sheet, rec) in head.items():
        base_entry = base.get(row_id)
        if base_entry is None or content_hash(base_entry[1]) != content_hash(rec):
            changed.add(row_id)
    return changed


def approving_review(repo: str, pr_number: int) -> tuple[str, str] | None:
    """(github_handle, YYYY-MM-DD) of the latest APPROVED review, or None."""
    result = subprocess.run(
        ["gh", "pr", "view", str(pr_number), "--repo", repo, "--json", "reviews"],
        capture_output=True, text=True, check=True,
    )
    reviews = json.loads(result.stdout)["reviews"]
    approvals = [r for r in reviews if r["state"] == "APPROVED"]
    if not approvals:
        return None
    latest = max(approvals, key=lambda r: r["submittedAt"])
    return latest["author"]["login"], latest["submittedAt"][:10]


def apply_json_stamps(write_dir: Path, row_ids: set[str], approver: str, approved_date: str) -> set[str]:
    """Stamp qualifying rows in data/json/*.json; return the IDs actually stamped."""
    stamped = set()
    for path in write_dir.glob("*.json"):
        if path.stem == "dictionary":
            continue
        data = load_sheet_json(path)
        if not data:
            continue
        changed = False
        for r in data.get("records", []):
            if r.get("id") in row_ids and not r.get("verified_by") and not r.get("last_verified"):
                r["verified_by"] = approver
                r["last_verified"] = approved_date
                stamped.add(r["id"])
                changed = True
        if changed:
            save_sheet_json(path, data)
    return stamped


def apply_xlsx_stamps(xlsx_path: Path, row_ids: set[str], approver: str, approved_date: str) -> None:
    if not row_ids:
        return
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)  # formulas preserved
    approved_date_cell = datetime.date.fromisoformat(approved_date)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {c.value: c.column_letter for c in ws[1]}
        if not {"ID", "Verified By", "Last Verified"} <= headers.keys():
            continue
        id_col, vb_col, lv_col = headers["ID"], headers["Verified By"], headers["Last Verified"]
        for row_idx in range(2, ws.max_row + 1):
            if ws[f"{id_col}{row_idx}"].value in row_ids:
                ws[f"{vb_col}{row_idx}"] = approver
                ws[f"{lv_col}{row_idx}"] = approved_date_cell
    wb.save(xlsx_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--base-dir", type=Path, required=True)
    parser.add_argument("--write-dir", type=Path, required=True, help="data/json inside the main checkout")
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--repo", required=True, help="owner/repo")
    parser.add_argument("--pr-number", type=int, required=True)
    args = parser.parse_args()

    approval = approving_review(args.repo, args.pr_number)
    if approval is None:
        print("No APPROVED review on this PR -- nothing to stamp.")
        return
    handle, approved_date = approval
    approver = f"@{handle}"

    changed_ids = find_changed_row_ids(args.base_dir, args.write_dir)
    if not changed_ids:
        print("No rows changed in this PR -- nothing to stamp.")
        return

    stamped = apply_json_stamps(args.write_dir, changed_ids, approver, approved_date)
    if not stamped:
        print("Changed rows already had verification -- nothing to stamp.")
        return

    apply_xlsx_stamps(args.xlsx, stamped, approver, approved_date)
    print(f"Stamped {len(stamped)} row(s), verified by {approver} on {approved_date}: {sorted(stamped)}")


if __name__ == "__main__":
    main()

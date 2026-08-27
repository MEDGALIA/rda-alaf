"""Convert the VT Radar workbook into JSON: one file per data sheet, plus a
hierarchical schema file built from the metadata tabs.

Bootstraps data/json/ from data/VANTAGE-Technology-Radar.xlsx, and is re-run
whenever the xlsx changes (by hand, or by the GitHub Action that reacts to an
uploaded xlsx in a PR).

The Dictionary/Vocabulary/Standards tabs are the schema, not records: they
become data/json/dictionary.json, with each column's controlled-vocabulary
terms (and their ontology mappings) nested underneath it. That schema also
drives the conversion -- a column declared `controlled_multi` becomes a real
JSON array, split on the separator the Dictionary tab specifies, rather than
the script hardcoding column names.

Row identity: each sheet's first column, `ID`, is the natural key used to
match a row across runs -- a system-generated value, independent of
`Resource Name`, so renaming a resource doesn't look like deleting one row
and adding an unrelated new one. Any row found with a blank `ID` cell gets
one assigned and written back into the xlsx before conversion starts (see
`assign_missing_ids()`). If an existing row's content changed since the last
run, its Verified By/Last Verified are cleared -- UNLESS a human filled in
both fields in that same xlsx edit, which is treated as a deliberate
re-verification and kept as-is: editing a row and re-confirming Verified By
in the same pass is exactly how a human is expected to (re-)verify it. A
brand-new row (no known prior baseline at all) always has its verification
trusted as-is, for the same reason.

Rows whose ID no longer appears in a sheet are checked against every other
data sheet before being reported: if the same ID now exists elsewhere (e.g.
a row moved from `Knowledgebase` to `Deprecated`, ID copied along with the
rest of the row), it's reported as a move, not a deletion. If the ID exists
nowhere else, it's reported as a deletion -- which today is only *reported*,
not blocked; enforcing "only an admin-merged PR may delete a row" is a
GitHub Action-side check (checklist item 19), not something this script can
decide on its own, since it has no notion of who's running it or approving
the PR.

`Added/Edited By` records who or what last touched a row's content --
self-declared in the cell by default, but overwritten with the value passed
via `--actor` on any row whose content actually changed in this run. Omit
`--actor` for a local run to leave the field as self-declared; a GitHub
Action passes its own trusted identity (e.g. `github.actor`) instead of
trusting free text.

`Last Updated` is the companion date: when this row's content was last
added or edited, auto-stamped with today's date on the same trigger as
`Added/Edited By` (content actually changed), unconditionally -- a date
carries no identity to distrust, so unlike `--actor` it doesn't need a
trusted caller to set it. This is *not* the resource's own date -- that's
`Publication/Version Date` (or `Version`, for a resource that has one),
which nothing here ever touches automatically.

Before any of this, the workbook's structure is validated against
`Dictionary`: if a data sheet's columns don't exactly match what `Dictionary`
declares for it (same headers, same order), conversion refuses to run at
all -- no partial/malformed JSON gets written.

When a row's verification is cleared, the corresponding xlsx cells are also
blanked (the script opens the workbook a second time, in write mode, only for
this). Without that, a cleared row's still-populated xlsx cell would silently
leak back into the JSON the next time that row's content changes again,
un-clearing something that was correctly cleared before -- this is otherwise
a normal read-only conversion, this is the one exception.

NOTE: adding or removing a column changes every row's content hash, so a
schema change makes this look like "every row was edited" and would clear all
verification for rows nobody actually touched. After any schema change, re-run
with --fresh, which rebuilds the baseline from the workbook instead of diffing
against the previous (differently-shaped) JSON.

Usage:
    python xlsx_to_json.py [--xlsx PATH] [--out-dir data/json] [--fresh]
"""

from __future__ import annotations

import argparse
import datetime
from pathlib import Path

import openpyxl

from radar_sync_common import (
    METADATA_SHEETS,
    cell_display_value,
    content_hash,
    iter_header,
    load_sheet_json,
    load_workbook_schema,
    new_id,
    row_is_empty,
    save_sheet_json,
    slugify,
    split_multi_value,
)

DEFAULT_XLSX = Path("data/VANTAGE-Technology-Radar.xlsx")
DEFAULT_OUT_DIR = Path("data/json")

VERIFICATION_FIELDS = ("verified_by", "last_verified")


def validate_structure(wb, schema: dict) -> None:
    """Raise if any data sheet's columns don't exactly match Dictionary.

    Runs before anything is written, so a malformed upload (wrong/missing/
    reordered/renamed columns) is rejected outright rather than partially
    converted.
    """
    problems = []
    for sheet_name in wb.sheetnames:
        if sheet_name in METADATA_SHEETS:
            continue
        expected = [c["name"] for c in sorted(schema["columns"], key=lambda c: c["position"]) if sheet_name in c["applies_to"]]
        actual = [h for _c, h in iter_header(wb[sheet_name])]
        if actual != expected:
            missing = [h for h in expected if h not in actual]
            extra = [h for h in actual if h not in expected]
            detail = []
            if missing:
                detail.append(f"missing: {missing}")
            if extra:
                detail.append(f"unexpected: {extra}")
            if not missing and not extra:
                detail.append(f"out of order: got {actual}, expected {expected}")
            problems.append(f"{sheet_name}: " + "; ".join(detail))
    if problems:
        raise ValueError(
            "xlsx structure doesn't match Dictionary -- refusing to convert:\n  "
            + "\n  ".join(problems)
        )


def assign_missing_ids(xlsx_path: Path) -> bool:
    """Fill any blank ID cell in a data sheet with a new one, saved in place.

    Runs before the main (data_only) read, so IDs are visible like any other
    pre-existing cell by the time conversion happens. Returns True if the
    xlsx was modified.
    """
    wb = openpyxl.load_workbook(xlsx_path)  # formulas preserved
    existing_ids: set[str] = set()
    blanks = []  # blank ID cells to fill in

    for sheet_name in wb.sheetnames:
        if sheet_name in METADATA_SHEETS:
            continue
        ws = wb[sheet_name]
        columns = list(iter_header(ws))
        if not columns or columns[0][1] != "ID":
            continue
        id_col = columns[0][0]
        for row_idx in range(2, ws.max_row + 1):
            if row_is_empty(ws, row_idx, columns):
                continue
            cell = ws[f"{id_col}{row_idx}"]
            if cell.value and str(cell.value).strip():
                existing_ids.add(str(cell.value).strip())
            else:
                blanks.append(cell)

    for cell in blanks:
        cell.value = new_id(existing_ids)

    if blanks:
        wb.save(xlsx_path)
    wb.close()
    return bool(blanks)


def sheet_to_records(ws, schema: dict) -> tuple[list[dict], list[str], dict[str, int]]:
    """Read a data sheet into records, using the schema to type each column.

    Also returns {natural_key: row_idx}, so the caller can write cleared
    verification back to the xlsx (see convert()) -- otherwise a cleared
    row's stale, never-blanked cell would leak back into the JSON the next
    time that row's content changes again.
    """
    columns = list(iter_header(ws))
    by_name = {c["name"]: c for c in schema["columns"]}

    records = []
    row_index_by_key = {}
    for row_idx in range(2, ws.max_row + 1):
        if row_is_empty(ws, row_idx, columns):
            continue
        record = {}
        for col_letter, header in columns:
            raw = ws[f"{col_letter}{row_idx}"].value
            spec = by_name.get(header)
            if spec and spec["value_type"] == "controlled_multi":
                record[slugify(header)] = split_multi_value(raw, spec["separator"] or ";")
            else:
                record[slugify(header)] = cell_display_value(raw)
        records.append(record)
        row_index_by_key[record[slugify(columns[0][1])]] = row_idx

    return records, [slugify(h) for _c, h in columns], row_index_by_key


def merge_with_previous(
    records: list[dict],
    previous: dict | None,
    key_field: str,
    actor: str | None = None,
    today: str | None = None,
) -> tuple[list[dict], list[str], list[str]]:
    """Clear verification on new/changed rows; report what changed.

    `actor`, when given, overwrites `added_edited_by` on any row whose
    content actually changed in this run (new row, or hash differs from the
    known baseline) -- an authoritative stamp of who/what made the edit,
    taking priority over whatever was self-declared in the cell. Left alone
    entirely when `actor` is None (e.g. a local run with no trusted identity
    to attribute changes to).

    `last_updated` -- the date THIS ROW was last added or edited, not the
    resource's own date (that's `publication_version_date`) -- is stamped
    with `today` on the same trigger, unconditionally: unlike `actor`, a
    date carries no identity to be untrusted, so there's no reason to gate
    it behind a trusted-caller check the way `actor` is. Defaults to the
    real current date; overridable for tests and for a deliberate historical
    backfill (never for normal runs).
    """
    if today is None:
        today = datetime.date.today().isoformat()
    prev_by_key = {r[key_field]: r for r in (previous or {}).get("records", [])} if previous else {}

    cleared, kept_keys = [], set()
    merged = []
    for record in records:
        key = record[key_field]
        kept_keys.add(key)

        prev_record = prev_by_key.get(key)
        new_hash = content_hash(record)
        content_changed = prev_record is None or prev_record.get("_content_hash") != new_hash
        if actor and content_changed and "added_edited_by" in record:
            record["added_edited_by"] = actor
        if content_changed and "last_updated" in record:
            record["last_updated"] = today

        if not any(f in record for f in VERIFICATION_FIELDS):
            record["_content_hash"] = new_hash
            merged.append(record)
            continue

        # content_hash() excludes verified_by/last_verified, so "content unchanged"
        # and "verification unchanged" are independent facts -- a human can edit
        # only the verification fields (e.g. filling in a fresh Last Verified with
        # no other change) without the hash moving at all. Comparing against the
        # value last recorded in the JSON (not merely checking "is it non-blank
        # now") is what distinguishes a fresh edit from a stale value just sitting
        # there unchanged from a previous, now-outdated verification.
        prev_verified_by = prev_record.get("verified_by", "") if prev_record else ""
        prev_last_verified = prev_record.get("last_verified", "") if prev_record else ""
        verified_by, last_verified = record.get("verified_by", ""), record.get("last_verified", "")
        verification_unchanged = verified_by == prev_verified_by and last_verified == prev_last_verified

        if not content_changed and verification_unchanged:
            # True no-op: nothing about this row changed since the last run.
            record["_content_hash"] = new_hash
            merged.append(record)
            continue

        # Either the content changed, or the human edited only the verification
        # fields directly (not content_changed but not verification_unchanged) --
        # both are edits that need a fresh-verification decision, so they share
        # the same check.
        fully_supplied = bool(verified_by) and bool(last_verified)
        freshly_touched = not verification_unchanged

        if fully_supplied and freshly_touched:
            # A human supplied (or changed) both fields in this same edit -- trust
            # it as a deliberate re-verification. This is the whole point of xlsx
            # being the human's verification channel.
            pass
        else:
            # Not a fresh, deliberate re-verification -- force fully blank.
            # Forcing blank even when it was already blank matters: without it, a
            # stale non-blank xlsx cell that was never physically cleared (see the
            # write-back step in convert()) would leak back in here and silently
            # un-clear a row that was correctly cleared before.
            was_verified = bool(verified_by or last_verified)
            record["verified_by"] = ""
            record["last_verified"] = ""
            if was_verified:
                cleared.append(key)
        record["_content_hash"] = new_hash
        merged.append(record)

    removed = sorted(k for k in prev_by_key if k not in kept_keys)
    return merged, cleared, removed


def convert(xlsx_path: Path, out_dir: Path, fresh: bool = False, actor: str | None = None) -> None:
    if assign_missing_ids(xlsx_path):
        print("assigned a new ID to every row that was missing one")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if fresh:
        print("--fresh: rebuilding the baseline from the workbook, ignoring existing JSON")

    schema = load_workbook_schema(wb)
    validate_structure(wb, schema)
    schema_path = out_dir / "dictionary.json"
    save_sheet_json(schema_path, schema)
    controlled = sum(1 for c in schema["columns"] if c["terms"])
    print(
        f"schema: {len(schema['columns'])} columns "
        f"({controlled} with vocabularies), {len(schema['standards'])} standards -> {schema_path}"
    )

    # (sheet_name, row_idx, field) for cells where the JSON says "blank" but
    # the xlsx cell itself isn't -- must be blanked too, or a later run that
    # re-reads the xlsx fresh will let the stale value leak back into the
    # JSON, silently un-clearing a row that was correctly cleared before.
    to_blank: list[tuple[str, int, str]] = []

    # (sheet_name, row_idx, value) for rows where --actor overwrote
    # added_edited_by in memory -- written back into the xlsx cell too, for
    # the same reason: the two representations must never be allowed to
    # drift apart.
    to_stamp: list[tuple[str, int, str]] = []

    # Per-sheet results, kept until every sheet's been processed -- a row
    # "removed" from one sheet can't be classified as moved vs. deleted until
    # we know what's currently present across *all* sheets, including ones
    # not processed yet in this loop.
    sheet_removed: dict[str, list[str]] = {}
    present_elsewhere: dict[str, tuple[str, str]] = {}  # id -> (sheet_name, resource_name)

    for sheet_name in wb.sheetnames:
        if sheet_name in METADATA_SHEETS:
            continue
        ws = wb[sheet_name]
        records, columns, row_index = sheet_to_records(ws, schema)
        if not columns:
            continue
        key_field = columns[0]

        out_path = out_dir / f"{slugify(sheet_name)}.json"
        previous = None if fresh else load_sheet_json(out_path)
        merged, cleared, removed = merge_with_previous(records, previous, key_field, actor=actor)

        header_letters = {slugify(h): c for c, h in iter_header(ws)}
        for record in merged:
            row_idx = row_index[record[key_field]]
            for field in VERIFICATION_FIELDS:
                if field not in record or record[field]:
                    continue
                letter = header_letters.get(field)
                if letter and ws[f"{letter}{row_idx}"].value not in (None, ""):
                    to_blank.append((sheet_name, row_idx, field))
            if actor and record.get("added_edited_by") == actor:
                letter = header_letters.get("added_edited_by")
                if letter and str(ws[f"{letter}{row_idx}"].value or "") != actor:
                    to_stamp.append((sheet_name, row_idx, actor))
            present_elsewhere[record[key_field]] = (sheet_name, record.get("resource_name", ""))

        save_sheet_json(out_path, {"sheet_name": sheet_name, "columns": columns, "records": merged})
        sheet_removed[sheet_name] = removed

        print(f"{sheet_name}: wrote {len(merged)} records to {out_path}")
        for key in cleared:
            print(f"  cleared verification: {key!r} (content changed)")

    # Now that every sheet's current contents are known, classify each
    # removed row as a move (same ID found in a different sheet) or a
    # deletion (found nowhere) -- see checklist item 14/15 in the sync plan.
    for sheet_name, removed in sheet_removed.items():
        for row_id in removed:
            if row_id in present_elsewhere:
                dest_sheet, dest_name = present_elsewhere[row_id]
                print(f"  moved from {sheet_name}: {dest_name!r} (id {row_id}) -> now in {dest_sheet!r}")
            else:
                print(
                    f"  DELETED from {sheet_name}, id {row_id} -- no match in any other sheet. "
                    "Not blocked by this script; requires an admin-merged PR (checklist item 19, not yet enforced)."
                )

    wb.close()

    if to_blank or to_stamp:
        wb_write = openpyxl.load_workbook(xlsx_path)  # formulas preserved
        for sheet_name, row_idx, field in to_blank:
            ws = wb_write[sheet_name]
            header_letters = {slugify(h): c for c, h in iter_header(ws)}
            cell = ws[f"{header_letters[field]}{row_idx}"]
            cell.value = None  # NOTE: cell(..., value=None) is a no-op; must assign .value directly
        for sheet_name, row_idx, value in to_stamp:
            ws = wb_write[sheet_name]
            header_letters = {slugify(h): c for c, h in iter_header(ws)}
            ws[f"{header_letters['added_edited_by']}{row_idx}"] = value
        wb_write.save(xlsx_path)
        if to_blank:
            print(f"blanked {len(to_blank)} stale verification cell(s) directly in the xlsx (kept in sync with the JSON)")
            for sheet_name, row_idx, field in to_blank:
                print(f"  {sheet_name} row {row_idx}: {field}")
        if to_stamp:
            print(f"stamped 'Added/Edited By' on {len(to_stamp)} changed row(s) directly in the xlsx")
            for sheet_name, row_idx, value in to_stamp:
                print(f"  {sheet_name} row {row_idx}: {value!r}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert the VT Radar xlsx into JSON.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    parser.add_argument(
        "--fresh", action="store_true",
        help="Rebuild the baseline from the workbook, ignoring existing JSON. Use after any "
             "column/schema change, which otherwise looks like every row was edited.",
    )
    parser.add_argument(
        "--actor", default=None,
        help="Trusted identity to stamp into 'Added/Edited By' on any row whose content "
             "changed in this run, overriding whatever was self-declared in the cell. "
             "E.g. a GitHub Action passing github.actor. Omit for a local run to leave "
             "the field as self-declared.",
    )
    args = parser.parse_args()

    convert(args.xlsx, args.out_dir, fresh=args.fresh, actor=args.actor)


if __name__ == "__main__":
    main()

"""Regenerate the VT Radar xlsx from data/json/*.json.

data/json/ is the source of truth; the xlsx is a generated artifact for
human consumption. This is the mirror of xlsx_to_json.py: it rebuilds the
**entire** workbook from scratch every run (every sheet, every row, every
metadata tab) rather than patching cells in place. It never reads the
previous xlsx. This is unrelated to the narrow write-back xlsx_to_json.py
does for exactly two stale verification cells -- that's an incidental step
inside the opposite (xlsx -> json) conversion, not a partial version of
this script.

Cosmetic details (column widths, frozen header row, bold headers) are sane
regenerated defaults, not a pixel-for-pixel clone of any hand-tweaked
formatting a curator may have applied to a previous xlsx.

URL hyperlinks and the Maturity Level / Topic Focus conditional-formatting
colours *are* regenerated, though -- they were present in the original
workbook (commit 55f1514) and are content-adjacent enough (a link you can
click, a status colour you can scan) that silently dropping them on every
regenerate is a real functional loss, not just cosmetics. Both are resolved
by column *name* against the schema, not a hardcoded letter, so they can't
drift out of alignment the way the original workbook's did after later
column inserts (see drafts/Color-coded-standards.md).

Usage:
    python json_to_xlsx.py [--json-dir data/json] [--out PATH]
"""

from __future__ import annotations

import argparse
import datetime
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from radar_sync_common import (
    CONTROLLED_TYPES,
    load_sheet_json,
)

DEFAULT_JSON_DIR = Path("data/json")
DEFAULT_OUT = Path("data/VANTAGE-Technology-Radar.xlsx")

# Preferred left-to-right tab order for data sheets; anything found on disk
# but not listed here is appended afterward, alphabetically.
DATA_SHEET_ORDER = ["knowledgebase", "deprecated", "sota_coding_agents_benchmarks"]

HEADER_FONT = Font(bold=True)
HYPERLINK_FONT = Font(color="FF0563C1", underline="single")

# Colours as they existed in the original workbook (commit 55f1514).
# "Adolescent" is deliberately dropped: not a real Maturity Level term (the
# vocabulary is Deprecated/Emerging/Mature/Research), a dead rule even then.
MATURITY_FILL = {
    "Emerging": "FFB7E1CD",
    "Mature": "FFA4C2F4",
    "Research": "FFFFE599",
    "Deprecated": "FFDD7E6B",
}
TOPIC_FOCUS_FILL = "FFB7E1CD"


def _autosize(ws: Worksheet, headers: list[str]) -> None:
    for i, header in enumerate(headers, start=1):
        width = max(10, min(60, len(header) + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for i, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    _autosize(ws, headers)


def _cell_value_for(value_type: str | None, separator: str | None, raw):
    """Convert a JSON field back into what a cell should hold."""
    if value_type == "controlled_multi":
        items = raw or []
        return (separator or ";").join(items) if items else None
    if not raw and raw != 0:
        return None
    if value_type == "date":
        try:
            return datetime.date.fromisoformat(str(raw))
        except ValueError:
            return raw  # not a full ISO date (e.g. a bare year) -- keep as text
    return raw


def _apply_url_hyperlinks(ws: Worksheet, keys: list[str], last_row: int) -> None:
    if "url" not in keys:
        return
    col = keys.index("url") + 1
    letter = get_column_letter(col)
    for row_idx in range(2, last_row + 1):
        cell = ws[f"{letter}{row_idx}"]
        if cell.value:
            cell.hyperlink = cell.value
            cell.font = HYPERLINK_FONT


def _apply_conditional_formatting(ws: Worksheet, keys: list[str], last_row: int) -> None:
    if last_row < 2:
        return

    if "maturity_level" in keys:
        letter = get_column_letter(keys.index("maturity_level") + 1)
        rng = f"{letter}2:{letter}{last_row}"
        for term, colour in MATURITY_FILL.items():
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=[f'"{term}"'], fill=_solid_fill(colour))
            )

    if "topic_focus" in keys:
        letter = get_column_letter(keys.index("topic_focus") + 1)
        rng = f"{letter}2:{letter}{last_row}"
        ws.conditional_formatting.add(
            rng, FormulaRule(formula=[f"LEN(TRIM({letter}2))>0"], fill=_solid_fill(TOPIC_FOCUS_FILL))
        )


def _solid_fill(rgb: str) -> PatternFill:
    return PatternFill(start_color=rgb, end_color=rgb, fill_type="solid")


def build_data_sheet(wb, sheet_json: dict, columns_by_key: dict[str, dict]) -> None:
    ws = wb.create_sheet(sheet_json["sheet_name"])
    keys = sheet_json["columns"]
    headers = [columns_by_key[k]["name"] if k in columns_by_key else k for k in keys]
    _write_header(ws, headers)

    for row_idx, record in enumerate(sheet_json["records"], start=2):
        for col_idx, key in enumerate(keys, start=1):
            spec = columns_by_key.get(key)
            value_type = spec["value_type"] if spec else None
            separator = spec["separator"] if spec else None
            ws.cell(row=row_idx, column=col_idx, value=_cell_value_for(value_type, separator, record.get(key)))

    last_row = len(sheet_json["records"]) + 1
    _apply_url_hyperlinks(ws, keys, last_row)
    _apply_conditional_formatting(ws, keys, last_row)


def build_dictionary_sheet(wb, schema: dict) -> None:
    ws = wb.create_sheet("Dictionary")
    headers = ["Position", "Column Name", "Applies To", "Value Type", "Separator", "Description"]
    _write_header(ws, headers)
    ws.column_dimensions["F"].width = 80

    for row_idx, col in enumerate(schema["columns"], start=2):
        ws.cell(row=row_idx, column=1, value=col["position"])
        ws.cell(row=row_idx, column=2, value=col["name"])
        ws.cell(row=row_idx, column=3, value="; ".join(col["applies_to"]))
        ws.cell(row=row_idx, column=4, value=col["value_type"])
        ws.cell(row=row_idx, column=5, value=col["separator"])
        ws.cell(row=row_idx, column=6, value=col["description"] or None)


def build_vocabulary_sheet(wb, schema: dict) -> None:
    ws = wb.create_sheet("Vocabulary")
    headers = ["Column Name", "Term", "Definition", "Ontology", "Ontology ID", "Ontology URL"]
    _write_header(ws, headers)
    ws.column_dimensions["C"].width = 80

    row_idx = 2
    for col in schema["columns"]:
        if col["value_type"] not in CONTROLLED_TYPES:
            continue
        for term in col["terms"]:
            ws.cell(row=row_idx, column=1, value=col["name"])
            ws.cell(row=row_idx, column=2, value=term["name"])
            ws.cell(row=row_idx, column=3, value=term["definition"] or None)
            ws.cell(row=row_idx, column=4, value=term["ontology"] or None)
            ws.cell(row=row_idx, column=5, value=term["ontology_id"] or None)
            ws.cell(row=row_idx, column=6, value=term["ontology_url"] or None)
            row_idx += 1


def build_standards_sheet(wb, schema: dict) -> None:
    ws = wb.create_sheet("Standards")
    headers = ["Standard ID", "Name", "Version", "URL", "Notes"]
    _write_header(ws, headers)
    ws.column_dimensions["E"].width = 90

    for row_idx, standard in enumerate(schema["standards"], start=2):
        ws.cell(row=row_idx, column=1, value=standard["id"])
        ws.cell(row=row_idx, column=2, value=standard["name"])
        ws.cell(row=row_idx, column=3, value=standard["version"] or None)
        ws.cell(row=row_idx, column=4, value=standard["url"] or None)
        ws.cell(row=row_idx, column=5, value=standard["notes"] or None)


def discover_data_sheet_files(json_dir: Path) -> list[Path]:
    files = [p for p in json_dir.glob("*.json") if p.stem != "dictionary"]

    def sort_key(p: Path):
        try:
            return (DATA_SHEET_ORDER.index(p.stem), p.stem)
        except ValueError:
            return (len(DATA_SHEET_ORDER), p.stem)

    return sorted(files, key=sort_key)


def convert(json_dir: Path, out_path: Path) -> None:
    schema = load_sheet_json(json_dir / "dictionary.json")
    if schema is None:
        raise FileNotFoundError(f"{json_dir / 'dictionary.json'} not found -- nothing to build from")
    columns_by_key = {c["key"]: c for c in schema["columns"]}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    for path in discover_data_sheet_files(json_dir):
        sheet_json = load_sheet_json(path)
        build_data_sheet(wb, sheet_json, columns_by_key)
        print(f"{sheet_json['sheet_name']}: wrote {len(sheet_json['records'])} rows")

    build_dictionary_sheet(wb, schema)
    build_vocabulary_sheet(wb, schema)
    build_standards_sheet(wb, schema)
    print(f"Dictionary: wrote {len(schema['columns'])} rows")
    print(f"Vocabulary: wrote {sum(len(c['terms']) for c in schema['columns'])} rows")
    print(f"Standards: wrote {len(schema['standards'])} rows")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"-> {out_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Regenerate the VT Radar xlsx from data/json/*.json.")
    parser.add_argument("--json-dir", type=Path, default=DEFAULT_JSON_DIR)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    convert(args.json_dir, args.out)


if __name__ == "__main__":
    main()

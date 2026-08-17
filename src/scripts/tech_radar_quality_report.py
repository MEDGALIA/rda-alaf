"""Read-only data-quality scan for the VANTAGE Technology Radar workbook.

Never opens the xlsx in write mode and never modifies any cell. Flags:
  - contiguous ranges of fully-empty rows
  - cells whose type/shape doesn't match what their column expects,
    based on simple heuristics keyed off the column header text

Output is a Markdown report; no "corrected" values are suggested, since
guessing at real WG records could be worse than leaving them alone.

Usage:
    python tech_radar_quality_report.py [--xlsx PATH] [--out drafts/workbook_errors.md]
"""

from __future__ import annotations

import argparse
import datetime
from pathlib import Path

import openpyxl

from radar_sync_common import iter_header, row_is_empty

DEFAULT_XLSX = Path("data/VANTAGE-Technology-Radar.xlsx")
DEFAULT_OUT = Path("drafts/workbook_errors.md")

NAME_LIKE_HEADERS = {"verified by", "source organization", "resource name"}
DATE_EXPECTED_HEADERS = {"last verified", "deprecation date"}
VERSION_HEADERS = {"version"}
URL_HEADERS = {"url"}


def is_date_cell(value) -> bool:
    return isinstance(value, (datetime.date, datetime.datetime))


def is_numeric_cell(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def check_anomaly(header: str, value) -> str | None:
    """Return a description of the anomaly, or None if the cell looks fine."""
    if value is None or str(value).strip() == "":
        return None
    header_lc = header.strip().lower()

    if header_lc in NAME_LIKE_HEADERS:
        if is_date_cell(value):
            return f"'{header}' holds a date, expected a name/organization/title"
        if is_numeric_cell(value):
            return f"'{header}' holds a number ({value!r}), expected a name/organization/title"
    elif header_lc in DATE_EXPECTED_HEADERS:
        if not is_date_cell(value):
            return f"'{header}' is not a real date cell (stored as {type(value).__name__}: {value!r})"
    elif header_lc in VERSION_HEADERS:
        if is_date_cell(value):
            return f"'{header}' holds a date ({value!r}), expected a version string"
    elif header_lc in URL_HEADERS:
        text = str(value).strip()
        if not (text.startswith("http://") or text.startswith("https://")):
            return f"'{header}' does not look like a URL: {value!r}"
    return None


def format_row_ranges(rows: list[int]) -> str:
    if not rows:
        return "none"
    ranges = []
    start = prev = rows[0]
    for r in rows[1:]:
        if r == prev + 1:
            prev = r
            continue
        ranges.append((start, prev))
        start = prev = r
    ranges.append((start, prev))
    return ", ".join(f"{a}" if a == b else f"{a}–{b}" for a, b in ranges)


def scan_sheet(ws) -> dict:
    columns = list(iter_header(ws))
    max_row = ws.max_row

    empty_rows = []
    populated_rows = 0
    anomalies = []  # (row, header, value, reason)

    for row_idx in range(2, max_row + 1):
        if row_is_empty(ws, row_idx, columns):
            empty_rows.append(row_idx)
            continue
        populated_rows += 1
        for col_letter, header in columns:
            value = ws[f"{col_letter}{row_idx}"].value
            reason = check_anomaly(header, value)
            if reason:
                anomalies.append((row_idx, header, value, reason))

    return {
        "columns": [h for _c, h in columns],
        "total_rows": max_row - 1,
        "empty_rows": empty_rows,
        "populated_rows": populated_rows,
        "anomalies": anomalies,
    }


def render_report(xlsx_path: Path, sheet_results: dict) -> str:
    lines = [
        "# VANTAGE Technology Radar — Workbook Data-Quality Report",
        "",
        f"Source: `{xlsx_path.as_posix()}`",
        "",
        "Read-only scan. No cell in the workbook was modified. Anomalies are flagged, not corrected — fixes are a human judgment call, made by hand in Excel.",
        "",
    ]
    for sheet_name, result in sheet_results.items():
        lines.append(f"## {sheet_name}")
        lines.append("")
        lines.append(f"- Total data rows: {result['total_rows']}")
        lines.append(f"- Empty rows: {len(result['empty_rows'])} (rows {format_row_ranges(result['empty_rows'])})")
        lines.append(f"- Populated rows: {result['populated_rows']}")
        lines.append(f"- Anomalous cells: {len(result['anomalies'])}")
        lines.append("")
        if result["anomalies"]:
            lines.append("| Row | Column | Value | Issue |")
            lines.append("|---|---|---|---|")
            for row_idx, header, value, reason in result["anomalies"]:
                lines.append(f"| {row_idx} | {header} | `{value!r}` | {reason} |")
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(description="Scan the Tech Radar workbook for data-quality issues.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    before_bytes = args.xlsx.read_bytes()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    sheet_results = {name: scan_sheet(wb[name]) for name in wb.sheetnames}
    wb.close()

    after_bytes = args.xlsx.read_bytes()
    assert before_bytes == after_bytes, "xlsx bytes changed during a read-only scan — this should never happen"

    report = render_report(args.xlsx, sheet_results)
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(report, encoding="utf-8", newline="\n")
    print(f"Wrote {args.out}")


if __name__ == "__main__":
    main()
